import os
from fastapi import FastAPI
from typing import List, Union
from joblib import Parallel, delayed
from pydantic import BaseModel # To serialize the data into JSON for API responses
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import joblib
from rtree import index
import rtree
from rdkit.Chem import rdMolDescriptors
import tmap as tm
from rdkit import Chem
import numpy as np
from rdkit.Chem import Draw
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image

app = FastAPI()

# Allow requests from the local frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:5501", # local setup
                   "http://localhost", # Docker container 
                   "http://localhost:80",
                   "http://localhost:3000" # Port where the frontend is running (docker)
                   # in production should be something like
                   # "https://chelombus.com",
                   ],  
    allow_credentials=True, # ALlows cookies/auth headers
    allow_methods=["*"],
    allow_headers=["*"], 
)

def export_smiles_to_excel(smiles_list: list,
                           output_excel: str="my_molecules.xlsx", 
                           image: bool=False):
    """
    Generates an Excel file with columns:
    SMILES_2D (Opt) |SMILES | Coordinate | Cluster_ID | TMAP_Link

    :param smiles_list: List of SMILES strings
    :param output_excel: Path to the output Excel file name

    The excel is a good way of going around the problem of plotting multiple points in the same map
    where several points can be mapped to the same cluster -meaning we lose the sense of how many of 
    our molecules are actually in the same cluster and how many are not. If multiple queries fall into 
    the same cluster then they will only appear as one point. I also tested changing the color based on 
    number of queried molecules per cluster but still we're missing which molecules are in which clusters
    
    This script contains also a 2D representation of the molecules, but I found that takes more time to 
    compute and you have to save and delete the images. This option can be done when `image=True`.
    """
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "SMILES_Clusters"

    # Headers, Column 1 contain images if image=True; if False, column is empty
    ws.cell(row=1, column=1, value="SMILES 2D structure")
    ws.cell(row=1, column=2, value="SMILES")
    ws.cell(row=1, column=3, value="Coordinate (x,y,z)")
    ws.cell(row=1, column=4, value="Cluster_ID")
    ws.cell(row=1, column=5, value="TMAP_Link")

    # Adjust column widths (optional, to ensure better readability)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 50

    row_idx = 2  # Start populating data at row 2
    i = 0
    df = pd.read_csv("/data/cluster_ranges.csv")
    pca_model = joblib.load("data/ipca_model.joblib")

    fps = mqn_parallel(smiles_list=smiles_list) 
    for smi in smiles_list:
        smi = smi.strip()
        if not smi:
            continue
        
        # 1) Find the cluster & coordinates from your existing function
        coordinate, cluster_id = find_cluster_from_smiles(pca_model=pca_model, cluster_ranges_dataframe=df, smiles=smi)
        if coordinate is None or cluster_id is None:
            # If there's no match, you might either skip or fill in "N/A"
            continue

        x, y, z = coordinate

        # 2) If image=True, create & insert 2D structure
        if image:
            mol = Chem.MolFromSmiles(smi)
            if mol:
                image_path = f"mol_{row_idx}.png"
                Draw.MolToFile(mol, image_path, size=(300, 300))
                
                # Insert the image into Excel
                img = Image(image_path)
                cell_location_for_image = f"A{row_idx}"  # Column A, row row_idx
                ws.add_image(img, cell_location_for_image)
                
                # Remove/cleanup the temporary PNG
                os.remove(image_path)

        # 3) Fill in the other columns
        # SMILES
        ws.cell(row=row_idx, column=2, value=smi)
        # Coordinates
        ws.cell(row=row_idx, column=3, value=f"({x:.2f}, {y:.2f}, {z:.2f})")
        # Cluster ID
        ws.cell(row=row_idx, column=4, value=cluster_id)
        # TMAP link
        tmap_link = f"http://localhost/generated_tmaps/cluster_{cluster_id}_TMAP.html"
        cell_for_link = ws.cell(row=row_idx, column=5, value="TMAP Link")
        cell_for_link.hyperlink = tmap_link
        cell_for_link.style = "Hyperlink"

        row_idx += 1
        i += 1
        print(f"\r Molecules done: {i}/{len(smiles_list)}", flush=True, end="")

    # Save the workbook
    wb.save(output_excel)
    print(f"Excel file saved: {output_excel}")

def calculate_mqn_fp(smiles: str) -> np.array:
    """Calculate MQN fingerprint for a single SMILES string.
    :param smiles: SMILES string for the molecule
    :return: np.array of fingerprint
    """
    try:
        fingerprint = rdMolDescriptors.MQNs_(Chem.MolFromSmiles(smiles))
        return np.array(fingerprint)
    except Exception as e:
        print(f"Error processing SMILES '{smiles}': {e}")
        return None
    
def mqn_parallel(smiles_list: list) -> list:
    """
    Calculate MQN fingerprints for a list of SMILES strings in parallel.
    
    :param smiles_list: List of SMILES strings.
    :return: List of numpy arrays (fingerprints), one for each SMILES.
    """
    fingerprints = Parallel(n_jobs=-1)(
        delayed(calculate_mqn_fp)(smiles) for smiles in smiles_list
    )
    return fingerprints


# Base Model for Coordinates 
class Coordinates(BaseModel):
    x: float
    y: float
    z: float

class ResponseModel(BaseModel):
    """
    Class to define the API response. Should contain a list of Coordinates objects
    under the coordinates key
    """
    coordinates: Union[Coordinates, List[Coordinates]]


def _get_coordinates_by_node_number(node_number) -> tuple:
    """
    Searches the CSV file for a row where 'cluster_name' contains 'cluster_<node_number>_TMAP.html'
    and returns the corresponding x, y, and z coordinates.
    
    :param csv_file: The path to the CSV file. This csv contains the coordinates (x,y,z) in the Primary TMAP
    for each cluster node. This way we can map our node_number with the actual location of that node in the 
    Primary TMAP HTML file.  
    :param node_number: The integer node number to search for.

    A tuple (x, y, z) if found, or None if not found.
    """
    # Load the CSV file
    cluster_coordinates_df= pd.read_csv('data/new_clusters_coordinates.csv')
    
    # Construct the target cluster name
    target_name = f"cluster_{node_number}_TMAP.html"
    
    # Search for the row with the matching cluster name
    matching_row = cluster_coordinates_df[cluster_coordinates_df['cluster_name'] == target_name]
    
    if not matching_row.empty:
        # Extract the x, y, z values
        x = matching_row.iloc[0]['x']
        y = matching_row.iloc[0]['y']
        z = matching_row.iloc[0]['z']
        return (x, y, z)
    else:
        return None  # Return None if no matching row is found
    

def find_cluster_from_smiles(cluster_ranges_df: pd.DataFrame, smiles: str):
    """
    Returns the (x, y, z) coordinate for the first matching cluster ID and the cluster_id
    found in `csv_file` for the input SMILES.
    :param csv_file: CSV file that includes the PCA ranges for each cluster. We find the cluster based on this.
    :param smiles: smiles string to find its cluster

    In order to get the PCA ranges from the database where we have all our data we can do it with the following query
    "SELECT cluster_id, MIN(PCA_1) AS min_PCA_1, MAX(PCA_1) AS max_PCA_1, MIN(PCA_2) AS min_PCA_2, MAX(PCA_2) 
    AS max_PCA_2, MIN(PCA_3) AS min_PCA_3, MAX(PCA_3) AS max_PCA_3 FROM clustered_enamine 
    GROUP BY cluster_id ORDER BY cluster_id ASC"
    as this can take a while we load it from a csv file with that data
    """
    # 1) Load the CSV file

    # 2) Load your trained PCA model
    pca = joblib.load("data/ipca_model.joblib")

    # 3) Compute PCA coordinates
    fingerprint = calculate_mqn_fp(smiles=smiles).reshape(1, -1)
    pca_coordinates = pca.transform(fingerprint).reshape(3, 1)

    # 4) Filter dataframe
    matching_clusters = cluster_ranges_df[
        (cluster_ranges_df['min_PCA_1'] <= np.float64(pca_coordinates[0])) & (cluster_ranges_df['max_PCA_1'] >= np.float64(pca_coordinates[0])) &
        (cluster_ranges_df['min_PCA_2'] <= np.float64(pca_coordinates[1])) & (cluster_ranges_df['max_PCA_2'] >= np.float64(pca_coordinates[1])) &
        (cluster_ranges_df['min_PCA_3'] <= np.float64(pca_coordinates[2])) & (cluster_ranges_df['max_PCA_3'] >= np.float64(pca_coordinates[2]))
    ]

    # 5) Get the list of matching cluster IDs (node_number)
    if matching_clusters.empty:
        return None

    cluster_id = matching_clusters['cluster_id'].tolist()

    # Just use the first matched node_number
    cluster_id= cluster_id[0]

    coordinate = _get_coordinates_by_node_number(cluster_id)
    # coordinate is either (x, y, z) or None
    
    return coordinate, cluster_id

@app.post("/api/find-cluster-from-jmse", response_model=ResponseModel)
async def find_cluster_from_jmse(smiles: dict) -> ResponseModel:
    """
    Function to find the cluster of a smiles that was pasted or drawn in the
    JMSE box in the frontend

    :param smiles: JSON body from the frontend containing the SMILES 
    string from the JMSE box
    """
    point = []
    df = pd.read_csv("data/cluster_ranges.csv")
    smiles_coordinate, cluster_id = find_cluster_from_smiles(
        cluster_ranges_df=df,
        smiles=smiles['smiles'], 
    )

    if smiles_coordinate:
        x, y, z = smiles_coordinate
        point.append(Coordinates(x=x, y=y, z=z))
    else: 
        # If there's no match, return None
        return ResponseModel(coordinates=None)

    return ResponseModel(coordinates=point)
    
@app.post("/api/find-cluster-from-file", response_model=ResponseModel)
async def find_cluster_from_file(request: dict) -> ResponseModel:
    """
    :param request: JSON body from the frontned.
    """
    points = []

    file_content = request.get("file_content", "")
    smiles_list= file_content.split("\n")

    for smiles in smiles_list:
        neighbor_smiles = smiles.strip()
        if not neighbor_smiles:
            continue  # Skip empty lines

        # find_cluster_from_smiles will return either (x, y, z) or None
        smiles_coordinate, cluster_id = find_cluster_from_smiles(
            csv_file="data/cluster_ranges.csv", 
            smiles=neighbor_smiles,
        )

        # If coordinate is found, append to our 'points' list
        if smiles_coordinate:
            x, y, z = smiles_coordinate
            points.append(Coordinates(x=x, y=y, z=z))
        else:
            # If there's no match, you could choose to do nothing
            return ResponseModel(coordinates=None)
    # Save file with results 
    export_smiles_to_excel(smiles_list=smiles_list, output_excel="cluster_search.xlsx")
    return ResponseModel(coordinates=points)